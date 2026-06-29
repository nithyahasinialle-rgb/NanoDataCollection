import csv, os, threading, time
import serial
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)

BAUD_RATE   = 115200
NUM_SAMPLES = 2000
END_MARKER  = "--- END OF DATA ---"
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
SAVE_DIR    = os.path.join(BASE_DIR, "Recordings")
ALL_SYMBOLS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + list("0123456789")
REPS        = 4
PORT        = "COM9"

# ── State ──────────────────────────────────────────────────────────────────────
# IMPORTANT: never call jsonify() or any Flask function while holding this lock.
# Copy values out first, then release the lock, then build the response.
lock  = threading.Lock()
state = {
    "recording":        False,
    "samples":          0,
    "live_data":        [],
    "selected_symbol":  "A",
    "current_attempt":  1,
    "volume":           "normal",
    "session_name":     "",
    "status":           "Ready – select a symbol and press Start",
    "log":              ["Connected on COM12"],
    "all_data":         {},   # { "A_1": [int, ...], ... }
    "completed":        [],   # ["A_1", "A_2", ...]
}


def log(msg):
    with lock:
        state["log"].append(msg)
        if len(state["log"]) > 300:
            state["log"].pop(0)
    print(msg)


def get_state_snapshot():
    """Copy all state values under lock, return plain dict safe to use outside lock."""
    with lock:
        return {
            "recording":       state["recording"],
            "samples":         state["samples"],
            "live_data":       list(state["live_data"]),
            "selected_symbol": state["selected_symbol"],
            "current_attempt": state["current_attempt"],
            "volume":          state["volume"],
            "session_name":    state["session_name"],
            "status":          state["status"],
            "log":             list(state["log"][-50:]),
            "all_data_keys":   list(state["all_data"].keys()),
            "completed":       list(state["completed"]),
        }


# ── Recording thread ───────────────────────────────────────────────────────────
def do_recording():
    # Grab everything we need then release lock immediately
    with lock:
        if state["recording"]:
            return
        state["recording"] = True
        state["samples"]   = 0
        state["live_data"] = []
        symbol = state["selected_symbol"]
        att    = state["current_attempt"]
        volume = state["volume"]

    log(f"Recording {symbol}_{att} [{volume}]")
    samples = []
    ser     = None

    try:
        ser = serial.Serial(PORT, BAUD_RATE, timeout=15)

        # DTR reset
        ser.dtr = False
        time.sleep(0.1)
        ser.dtr = True
        log("DTR reset sent")
        time.sleep(2.5)
        ser.reset_input_buffer()

        # Handshake
        ser.write(b"START\n")
        log("Sent START – waiting for READY...")
        deadline = time.time() + 10
        while True:
            if time.time() > deadline:
                raise TimeoutError("Arduino did not reply READY within 10s")
            line = ser.readline().decode("utf-8", errors="ignore").strip()
            if line == "READY":
                log("Arduino READY – sampling begins")
                break

        # Collect samples
        deadline = time.time() + 30
        while True:
            if time.time() > deadline:
                raise TimeoutError("Data collection timed out after 30s")
            raw = ser.readline()
            if not raw:
                continue
            line = raw.decode("utf-8", errors="ignore").strip()
            if line == END_MARKER:
                log(f"End marker received – {len(samples)} samples")
                break
            try:
                samples.append(int(line))
                # Update live_data every 25 samples (not every sample – avoids lock hammering)
                if len(samples) % 25 == 0:
                    with lock:
                        state["samples"]   = len(samples)
                        state["live_data"] = samples[:]
            except ValueError:
                pass  # skip non-numeric lines (bootloader noise etc.)

        if len(samples) < NUM_SAMPLES:
            raise ValueError(f"Incomplete: only {len(samples)} of {NUM_SAMPLES} samples received")

        # Save CSV
        folder = os.path.join(SAVE_DIR, symbol)
        os.makedirs(folder, exist_ok=True)
        path = os.path.join(folder, f"{symbol}_{att}.csv")
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Sample"])
            for v in samples:
                w.writerow([v])
        log(f"Saved → {path}")

        key = f"{symbol}_{att}"
        with lock:
            state["all_data"][key] = samples[:]
            if key not in state["completed"]:
                state["completed"].append(key)
            state["live_data"]     = samples[:]
            state["samples"]       = len(samples)
            state["status"]        = f"✓ Saved {key}  ({len(samples)} samples)"
            # Auto-advance attempt counter
            if state["current_attempt"] < REPS:
                state["current_attempt"] += 1
            else:
                state["status"] = f"✓ All 4 attempts for '{symbol}' done! Pick next symbol."

    except Exception as e:
        log(f"ERROR: {e}")
        with lock:
            state["status"] = f"Error: {e}"

    finally:
        # ALWAYS close the port and clear the recording flag
        if ser is not None:
            try:
                ser.close()
            except Exception:
                pass
        with lock:
            state["recording"] = False
        log("Recording thread finished – ready")


# ── Excel export ───────────────────────────────────────────────────────────────
def build_excel(filename):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Copy data out before doing any heavy work
    with lock:
        all_data = {k: list(v) for k, v in state["all_data"].items()}
        volume   = state["volume"]

    if not all_data:
        raise ValueError("No data recorded yet")

    wb     = Workbook()
    wb.remove(wb.active)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill  = PatternFill("solid", start_color="2D2D2D", end_color="2D2D2D")
    afill  = PatternFill("solid", start_color="F7F7F7", end_color="F7F7F7")

    # Group keys by symbol, sorted
    sym_map = {}
    for key in sorted(all_data.keys()):
        sym = key.split("_")[0]
        sym_map.setdefault(sym, []).append(key)

    for sym, keys in sym_map.items():
        ws = wb.create_sheet(title=sym)

        # Column headers
        for ci, key in enumerate(keys, 1):
            att = key.split("_")[1]
            cl  = get_column_letter(ci)
            c   = ws.cell(row=1, column=ci, value=f"Attempt {att}")
            c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill      = hfill
            c.alignment = Alignment(horizontal="center")
            c.border    = border
            ws.column_dimensions[cl].width = 14

        # Data rows
        max_len = max(len(all_data[k]) for k in keys)
        for ri in range(max_len):
            for ci, key in enumerate(keys, 1):
                d   = all_data[key]
                val = d[ri] if ri < len(d) else None
                c   = ws.cell(row=ri + 2, column=ci, value=val)
                c.font   = Font(name="Arial", size=9)
                c.border = border
                if ri % 2 == 0:
                    c.fill = afill

        # One line chart per attempt, placed below the data
        chart_row = max_len + 4
        for ci, key in enumerate(keys, 1):
            att  = key.split("_")[1]
            data = all_data[key]
            ch   = LineChart()
            ch.title        = f"{sym} – Attempt {att} [{volume}]"
            ch.style        = 10
            ch.y_axis.title = "Sensor Value"
            ch.x_axis.title = "Sample #"
            ch.height       = 10
            ch.width        = 20
            ref = Reference(ws, min_col=ci, max_col=ci,
                            min_row=2, max_row=len(data) + 1)
            ch.add_data(ref, titles_from_data=False)
            from openpyxl.chart.series import SeriesLabel
            ch.series[0].title       = SeriesLabel(v=f"Attempt {att}")
            ch.series[0].graphicalProperties.line.solidFill = "4F81BD"
            ch.series[0].graphicalProperties.line.width      = 15000
            ws.add_chart(ch, f"{get_column_letter((ci - 1) * 4 + 1)}{chart_row}")

    # Summary sheet at front
    ws_s = wb.create_sheet(title="Summary", index=0)
    for r, (a, b) in enumerate([
        ("Session",    filename),
        ("Volume",     volume.capitalize()),
        ("Symbols",    len(sym_map)),
        ("Recordings", len(all_data)),
    ], 1):
        ws_s.cell(row=r, column=1, value=a).font = Font(bold=True, name="Arial", size=10)
        ws_s.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws_s.column_dimensions["A"].width = 22
    ws_s.column_dimensions["B"].width = 28

    out = os.path.join(BASE_DIR, f"{filename}.xlsx")
    wb.save(out)
    log(f"Excel saved → {out}")
    return out


# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    # IMPORTANT: snapshot first, then build response OUTSIDE the lock
    snap = get_state_snapshot()
    return jsonify(snap)


@app.route("/api/select_symbol", methods=["POST"])
def api_select_symbol():
    d   = request.json or {}
    sym = str(d.get("symbol", "")).upper().strip()
    if sym not in ALL_SYMBOLS:
        return jsonify({"ok": False, "error": "Invalid symbol"})
    with lock:
        if state["recording"]:
            return jsonify({"ok": False, "error": "Recording in progress"})
        state["selected_symbol"] = sym
        state["current_attempt"] = 1
        state["live_data"]       = []
        state["samples"]         = 0
        # Find first incomplete attempt for this symbol
        for a in range(1, REPS + 1):
            if f"{sym}_{a}" not in state["all_data"]:
                state["current_attempt"] = a
                break
        else:
            state["current_attempt"] = REPS  # all done, stay on last
        state["status"] = f"Selected {sym} – attempt {state['current_attempt']}"
    return jsonify({"ok": True})


@app.route("/api/select_attempt", methods=["POST"])
def api_select_attempt():
    d   = request.json or {}
    att = d.get("attempt", 1)
    try:
        att = int(att)
    except (TypeError, ValueError):
        att = 1
    if not 1 <= att <= REPS:
        return jsonify({"ok": False, "error": "Attempt must be 1-4"})
    with lock:
        if state["recording"]:
            return jsonify({"ok": False, "error": "Recording in progress"})
        state["current_attempt"] = att
        sym = state["selected_symbol"]
        key = f"{sym}_{att}"
        if key in state["all_data"]:
            state["live_data"] = state["all_data"][key][:]
            state["samples"]   = len(state["live_data"])
        else:
            state["live_data"] = []
            state["samples"]   = 0
        state["status"] = f"Ready – {sym}_{att}"
    return jsonify({"ok": True, "attempt": att})


@app.route("/api/set_volume", methods=["POST"])
def api_set_volume():
    d   = request.json or {}
    vol = d.get("volume", "normal")
    if vol in ("whisper", "normal", "loud"):
        with lock:
            state["volume"] = vol
    return jsonify({"ok": True})


@app.route("/api/set_session", methods=["POST"])
def api_set_session():
    d = request.json or {}
    with lock:
        state["session_name"] = d.get("session_name", "").strip()
    return jsonify({"ok": True})


@app.route("/api/start", methods=["POST"])
def api_start():
    with lock:
        already = state["recording"]
        sym     = state["selected_symbol"]
    if already:
        return jsonify({"ok": False, "error": "Already recording"})
    if not sym:
        return jsonify({"ok": False, "error": "No symbol selected"})
    threading.Thread(target=do_recording, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/skip", methods=["POST"])
def api_skip():
    """Clear current attempt so it can be re-recorded."""
    with lock:
        if state["recording"]:
            return jsonify({"ok": False, "error": "Recording in progress"})
        sym = state["selected_symbol"]
        att = state["current_attempt"]
        key = f"{sym}_{att}"
        state["all_data"].pop(key, None)
        if key in state["completed"]:
            state["completed"].remove(key)
        state["live_data"] = []
        state["samples"]   = 0
        state["status"]    = f"Cleared {key} – ready to re-record"
    log(f"Cleared {key} for re-record")
    return jsonify({"ok": True})


@app.route("/api/next_attempt", methods=["POST"])
def api_next_attempt():
    with lock:
        if state["recording"]:
            return jsonify({"ok": False, "error": "Recording in progress"})
        att = (state["current_attempt"] % REPS) + 1   # 1→2→3→4→1
        state["current_attempt"] = att
        sym = state["selected_symbol"]
        key = f"{sym}_{att}"
        if key in state["all_data"]:
            state["live_data"] = state["all_data"][key][:]
            state["samples"]   = len(state["live_data"])
        else:
            state["live_data"] = []
            state["samples"]   = 0
        state["status"] = f"Ready – {sym}_{att}"
    return jsonify({"ok": True, "attempt": att})


@app.route("/api/clear", methods=["POST"])
def api_clear():
    with lock:
        if state["recording"]:
            return jsonify({"ok": False, "error": "Cannot clear while recording"})
        state["all_data"]        = {}
        state["completed"]       = []
        state["live_data"]       = []
        state["samples"]         = 0
        state["current_attempt"] = 1
        state["status"]          = "All data cleared – ready"
    log("All data cleared")
    return jsonify({"ok": True})


@app.route("/api/attempt_data")
def api_attempt_data():
    """Return stored samples for one symbol+attempt (used by chart loader)."""
    sym = request.args.get("symbol", "").upper().strip()
    att = request.args.get("attempt", "1").strip()
    key = f"{sym}_{att}"
    with lock:
        data = list(state["all_data"].get(key, []))
    return jsonify({"key": key, "data": data})


@app.route("/api/export", methods=["POST"])
def api_export():
    d = request.json or {}
    # Read name outside the heavy work
    with lock:
        fallback = state["session_name"]
    name = d.get("filename", "").strip() or fallback or "session"
    name = "".join(c for c in name if c.isalnum() or c in " _-").strip() or "session"
    try:
        path = build_excel(name)
        fname = os.path.basename(path)
        return send_file(path, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        log(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    os.makedirs(SAVE_DIR, exist_ok=True)
    print("Open  http://localhost:5000")
    # use_reloader=False prevents the reloader from spawning a second process
    # that would fight over the serial port
    app.run(host="127.0.0.1", port=5000, debug=False,
            threaded=True, use_reloader=False)